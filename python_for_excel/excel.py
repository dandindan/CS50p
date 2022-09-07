import openpyxl
print("hello")
# Load the file(change the Path to where the file is stored)
book  = openpyxl.load_workbook(r'/Users/maozlahav/Desktop/marks.xlsx')
sheet_results = book["Results"]


# Print rows 1 to 5 and columns 3 to 6.
for row in sheet_results.iter_rows(min_row=1,max_row=5,min_col=3,max_col=6,values_only=True):
    print(row)

print("cell 4,4 =" , sheet_results.cell(row=4,column=4).value)

cell_D8 = sheet_results["D8"]
print(cell_D8.value, cell_D8.row, cell_D8.column, cell_D8.data_type)
print(dir(cell_D8))
