import openpyxl

wb = openpyxl.load_workbook('store.xlsx')

print(wb.sheetnames)

#create new sheet named Maoz
# wb.create_sheet(title='Maoz') 
print(wb.worksheets)

for sheet in wb:
    print(sheet.title)

# save worksheet
#wb.save('store.xlsx')